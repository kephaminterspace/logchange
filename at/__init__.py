# -*- coding: utf-8 -*-
import sys
reload(sys)
sys.setdefaultencoding("utf-8")
from flask import Flask, Blueprint, request, render_template, url_for, send_from_directory
import logging
import datetime
import requests
import urllib
import json
import uuid
import os
from openpyxl import load_workbook, Workbook
from pyexcel_xlsx import get_data
import csv

from .forms import BankForm
from .utils import *

logger = logging.getLogger(__name__)
app = Flask(__name__, template_folder="templates", static_folder="statics")
#app.config.from_object("at.config")
app.secret_key = 'citibankfajdslkf'

@app.route('/robots.txt')
@app.route('/sitemap.xml')
def static_from_root():
    return send_from_directory(app.static_folder, request.path[1:])

@app.route("/", methods=['POST', 'GET'])
def index():
    form = BankForm()

    PARTH_LOCHANGE = "/media/mebao/DATA_UBUNTU/LAND_PAGE/LOG_CHANGE/logchange/at/statics"
    if not os.path.exists('/media/mebao'):
        PARTH_LOCHANGE="/opt/landingpages/logchange/at/statics"

    if not form.day_logchange.data:
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        form.day_logchange.data = lastMonth.strftime("%d/%m/%Y")

    if not form.folder_logchange.data:
        form.folder_logchange.data = "logchange"

    total_amount_Approved = 0
    quantity_Approved = 0
    quantity_Hold = 0
    quantity_Rejected = 0
    quantity_UpdateAmount = 0


    first_Title = ["ID", "TransID", "Date", "ProductID", "Quantity", "Sales Amount", "Reward Amount",
                   "Identifies", "Session ID", "device", "Check Date", "Result Target", "Creative",
                   "Partner Site Name", "Fee ID", "Fee Amount", "DeadLine", "Update Status"]

    updateamount_Title = ["ID", "TransID", "Date", "ProductID", "Quantity", "Update Sales Amount", "Reward Amount",
                          "Identifies",
                          "Session ID", "device", "Check Date", "Result Target", "Creative", "Partner Site Name",
                          "Fee ID",
                          "Fee Amount", "DeadLine", "Target"]

    data_Approved = []
    data_Hold = []
    data_Rejected = []
    data_UpdateAmount = []

    if request.method == 'POST':
        file_obj = request.files['datafile']
        wb = load_workbook(file_obj)
        sheets = wb.worksheets

        day_logchange = form.day_logchange.data
        folder_logchange = form.folder_logchange.data

        parth_File_Approved = PARTH_LOCHANGE + '/' + folder_logchange+'/MassApproved.csv'
        parth_File_UpdateAmount = PARTH_LOCHANGE + '/' + folder_logchange + '/UpdateAmount.csv'

        if len(sheets) > 0:
            if not os.path.exists(PARTH_LOCHANGE + '/' + folder_logchange):
                os.makedirs(PARTH_LOCHANGE + '/' + folder_logchange)

            if not os.path.exists(parth_File_Approved):
                os.mknod(parth_File_Approved)
            Approved = open(parth_File_Approved, "wb")
            writer_Approved = csv.writer(Approved, delimiter=',', quoting=csv.QUOTE_ALL)


            if not os.path.exists(parth_File_UpdateAmount):
                os.mknod(parth_File_UpdateAmount)
            UpdateAmount = open(parth_File_UpdateAmount, "wb")
            writer_UpdateAmount = csv.writer(UpdateAmount, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)

            writer_Approved.writerow(first_Title)
            writer_UpdateAmount.writerow(updateamount_Title)

            sheet = sheets[0]
            for row in sheet.iter_rows():
                if row[0].value != None and row[0].value !='':
                    item = None
                    if row[17].value.lower() == "approved" or row[17].value.lower() == "approve":
                        item = [row[0].value, None, None, None, None, None, None, None, None, None, day_logchange, None,
                                None, None, None, None, None, "Approved"]
                        quantity_Approved = quantity_Approved + int(row[4].value)
                        total_amount_Approved = total_amount_Approved + float(row[5].value)
                        if len(row)>20:
                            if row[20].value != '' and row[20].value != '-' and row[20].value != None and row[20].value != 0:
                                item1 = [row[0].value, None, None, None, None, round(float(row[20].value),2), None, None, None, None, day_logchange,
                                        None, None, None, None, None, None, "Approved"]
                                writer_UpdateAmount.writerow(item1)
                                quantity_UpdateAmount = quantity_UpdateAmount + int(row[4].value)
                                data_UpdateAmount.append(item1)

                    if row[17].value.lower() == "hold" or row[17].value.lower() == "New":
                        item = [row[0].value, None, None, None, None, None, None, None, None, None, day_logchange, None, None, None,
                                None, None, None, "Hold"]
                        quantity_Hold = quantity_Hold + int(row[4].value)


                    if row[17].value.lower() == "reject" or row[17].value.lower() == "rejecte" or row[17].value.lower() == "rejected":
                        item = [row[0].value, None, None, None, None, None, None, None, None, None, day_logchange, None,
                                None, None, None, None, None, "Rejected"]
                        quantity_Rejected = quantity_Rejected + int(row[4].value)

                    if item:
                        writer_Approved.writerow(item)
                        data_Approved.append(item)

            Approved.close()
            UpdateAmount.close()

    return render_template('index.html',
                           form = form,
                           PARTH_LOCHANGE=PARTH_LOCHANGE ,
                           folder_logchange=form.folder_logchange.data,
                           first_Title=first_Title,
                           updateamount_Title=updateamount_Title,

                           total_amount_Approved=total_amount_Approved,

                           quantity_Approved=quantity_Approved,
                           quantity_Hold=quantity_Hold,
                           quantity_Rejected=quantity_Rejected,
                           quantity_UpdateAmount=quantity_UpdateAmount,

                           data_Approved=data_Approved,
                           data_UpdateAmount=data_UpdateAmount
                           )
